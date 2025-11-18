#  11. Stwórz program który przyjmie w parametrze ścieżkę do dowolnego pliku (CSV lub Excel - jaki wolicie), który będzie zawierał dane tabelaryczne.
#    W pliku pierwszy wiersz będzie zawierał nazwy kolumn a pozostałe wiersze dane.
#    Ilość kolum i wierszy może być dowolna. Program ma narysować tabelę z danymi, analogicznie do wcześniejszego zadania na rysowanie tabeli.
#    Pamiętajmy by wydzielać części reużywalne do oddzielnych funkcji/modułów (np.: odczyt danych, przygotowanie danych, rysowanie tabeli).
#    Przykład:
#    +------------+------------+------------+
#    | klucz1     | klucz 2    | klucz 3    |
#    +------------+------------+------------+
#    | row 1 col1 | row 1 col2 | row 1 col3 |
#    +------------+------------+------------+
#    | row 2 col1 | row 2 col2 | row 2 col3 |
#    +------------+------------+------------+


import sys
import csv
from openpyxl import load_workbook
from pathlib import Path

# ===========================================================================================
# ---------- Funkcje wczytujące dane z pliku (CSV lub Excel)----------
# ===========================================================================================

def read_tabular_file(path_str: str):
    """
    Wczytuje dane tabelaryczne z pliku CSV lub Excel.
    Zwraca:
      - headers: list[str]   -> nazwy kolumn (pierwszy wiersz)
      - rows: list[list[str]] -> dane, każda lista to jeden wiersz
    """
    path = Path(path_str)
    if not path.exists():
        raise FileNotFoundError(f"Plik nie istnieje: {path}")

    suffix = path.suffix.lower()

    if suffix == ".csv":
        return read_csv_file(path)
    elif suffix in (".xlsx", ".xls"):
        return read_excel_file(path)
    else:
        raise ValueError(f"Nieobsługiwany typ pliku: {suffix}. Użyj CSV lub Excela.")


def read_csv_file(path: Path):
    """
    Wczytuje dane z pliku CSV.
    Założenia:
      - pierwszy wiersz to nagłówek,
      - separator kolumn to przecinek,
      - separator wierszy to znak nowej linii
    """
    headers = None
    rows = []

    with path.open("r", encoding="utf-8", newline="") as f: # newline="" - wyłączamy automatyczne "tłumaczenie" znaków końca linii - Python sam nie zmodyfikuje tych znaków
        reader = csv.reader(f) # iterator do czytania pliku linia po linii, rozdzieli wiersz po przecinkach; można doddać np. delimiter=';'

        first = True # czy to pierwszy wiersz (header)

        for row in reader:
            if first:
                headers = row
                first = False
            else:
                rows.append(row)

    # Jeśli plik był pusty, headers będzie dalej None
    if headers is None:
        print("Pusty plik!")
        return [], []  # zwracamy puste listy

    return headers, rows


def read_excel_file(path: Path):
    """
    Wczytuje dane z pliku Excel (.xlsx) używając openpyxl.
    """
    try:
        wb = load_workbook(filename=path, read_only=True)
    except Exception as e:
        raise IOError(f"Nie można otworzyć pliku Excel: {e}")

    # Domyślnie bierzemy aktywny (zwykle pierwszy) arkusz
    sheet = wb.active # możemy wybrać też po naziwe, np. sheet = wb['Arkusz1']


    rows_iter = sheet.iter_rows(values_only=True) # zwracamy iterator do przechodzenia wiersz po wierszu, zwracając tylko wartości komórek zamiast obiektów Cell, np. <Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>

    # Można użyć tego samego co w funkcji do otwierania CSV, ale dla celów naukowych próbujemy inaczej :)

    # Próbujemy pobrać pierwszy wiersz (header))
    try:
        first_row = next(rows_iter) # bierzemy wartości z pierwszego wiersza
    except StopIteration:
        # Jeśli nie ma nawet jednego wiersza (plik pusty), to zwracamy puste dane
        print("Pusty plik")
        return [], []

    # Przetwarzamy nagłówki — zamieniamy każde pole na tekst
    headers = []
    for cell in first_row:
        if cell is None:
            headers.append("")  # jeśli komórka była pusta, to dajemy pusty tekst
        else:
            headers.append(str(cell))  # inaczej zamieniamy wartość na tekst

    # przetwarzamy resztę wierszy (dane)
    rows = []  # tu będziemy przechowywać dane
    for row in rows_iter:
        current_row_list = []  # tu zapisujemy jedną linijkę danych

        for cell in row:
            if cell is None:
                current_row_list.append("")
            else:
                current_row_list.append(str(cell))

        rows.append(current_row_list)  # dodajemy wiersz do listy danych

    # Na koniec zwracamy nagłówki oraz wszystkie wiersze
    return headers, rows


# ===========================================================================================
# ----------  Funkcje przygotowujące dane do rysowania tabeli ----------
# ===========================================================================================

def normalize_rows(headers, rows):
    """
    Upewnia się, że każdy wiersz ma taką samą liczbę kolumn co nagłówek.
    Jeśli brakuje wartości – uzupełnia pustym stringiem (żeby nie było IndexErrorów)
    Jeśli jest za dużo to ucinamy nadmiar
    """
    num_cols = len(headers)
    normalized = []

    for row in rows:
        # Jeśli wiersz jest krótszy – dopełniamy pustymi stringami
        if len(row) < num_cols:
            row = list(row) + [""] * (num_cols - len(row))
        # Jeśli jest dłuższy – obcinamy nadmiar (rzadki przypadek)
        elif len(row) > num_cols:
            row = row[:num_cols]
        normalized.append(row)

    return normalized


def compute_different_column_widths(headers, rows):
    """
    Liczy szerokość każdej kolumny osobno na podstawie:
    - długości nagłówka,
    - długości wartości w wierszach.
    Nie używana dalej w programie, zrobiona dla celów naukowych :)
    """
    num_cols = len(headers) # liczba kolumn
    widths = [0] * num_cols

    for i in range(num_cols):
        # zaczynamy od długości nagłówka
        max_width = len(str(headers[i]))
        # porównujemy z każdą wartością w danej kolumnie
        for row in rows:
            cell_len = len(str(row[i]))
            if cell_len > max_width:
                max_width = cell_len
        widths[i] = max_width

    return widths


def compute_column_widths(headers, rows, max_col_width: int=30):
    """
    Zwraca listę szerokości kolumn — jednakową dla każdej kolumny.
    Szerokość zależy od:
    - długości najdłuższego nagłówka lub wartości w danych,
    - ale nie przekracza 'max_col_width'.
    """
    num_cols = len(headers) # liczba kolumn

    # Lista liczby znaków w kolumnach
    lengths = []

    # Sprawdzamy liczbę znaków w nagłówkach i dodajemy do lenghts
    for h in headers:
        text = str(h)
        lengths.append(len(text))

    # Sprawdzamy liczbę znaków w danych
    for row in rows: # lecimy po każdym wierszu
        for cell in row: # lecimy po każdej kolumnie (komórce)
            text = str(cell)
            lengths.append(len(text))

    # Sprawdzamy maksymalną długość
    if lengths:
        uniform_width = min(max_col_width, max(lengths)) # min - bierzemy mniejszą z dwóch liczb (albo 30 albo max z lengths)
    else:
        uniform_width = max_col_width  # jeśli brak danych, przyjmujemy defaultową wartość

    # Zwracamy jednakową szerokość dla każdej kolumny - budujemy listę kolum, np. [2] * 3 -> [2,2,2]
    return [uniform_width] * num_cols # robimy listę żeby było uniwersalnie i w dalszych funkcjach dało się też użyć funkcji z różnymi szerokościami kolumn


def truncate_to_width(value, width: int) -> str:
    """
    Zamienia wartość na tekst i dopasowuje do podanej szerokości kolumny.
    - jeśli tekst jest dłuższy niż width obcina go i dodaje '...' na końcu
    - jeśli krótszy dopełnia spacjami z prawej (wyrównanie do lewej).
    """
    text = str(value)

    if len(text) > width:
        # Jeśli width jest większe niż width - 3, to zarezerwujemy miejsce na '...'
        if width > 3:
            text = text[: width - 3] + "..."
        else:
            text = text[:width] # na wypadek zmiany max_col_width na wartość =< 3

    # Dopełniamy spacjami, żeby kolumna była równa
    return text.ljust(width)


# ===========================================================================================
# ---------- Funkcje rysujące tabelę----------
# ===========================================================================================

def draw_horizontal_border(col_widths) -> str:
    """
    Buduje poziomą linię tabeli
    """
    parts = [] # tu dodajemy fragmenty linii
    for width in col_widths:
        parts.append("+" + "-" * (width + 2)) # width + 2 -> jedna spacja z lewej, jedna z prawej w komórce
    parts.append("+") # "+" na końcu linii
    return "".join(parts) # łączymy wszystkie części w jeden string


def draw_row(cells, col_widths) -> str:
    """
    Rysuje pojedynczy wiersz tabeli
    Używa truncate_to_width, żeby nie „wyjeżdżać” poza kolumnę.
    """
    parts = [] # lista, która będzie zawierała kolejne części wiersza
    # zip paruje elementy dwóch list - my parujemy wartość komórki z szerokością kolumny
    for cell, width in zip(cells, col_widths): # cells - lista wartości z danego wiersza
        cell_text = truncate_to_width(cell, width) # zamieniamy wartość na tekst, jak trzeba to przycinamy
        parts.append("| " + cell_text + " ")
    parts.append("|")
    return "".join(parts) # sklejamy wszystkie części w jednego stringa


def draw_table(headers, rows, max_col_width: int = 30):
    """
    Główna funkcja rysująca tabelę:
      - normalizuje wiersze,
      - oblicza szerokości kolumn,
      - rysuje ramki, nagłówki i wiersze danych.
    """
    if not headers:
        print("Brak nagłówków – tabela nie może zostać narysowana.")
        return None

    # Upewniamy się, że każdy wiersz ma tyle samo kolumn co nagłówek
    rows = normalize_rows(headers, rows)

    # Obliczamy szerokość kolumn
    col_widths = compute_column_widths(headers, rows, max_col_width=max_col_width)

    border = draw_horizontal_border(col_widths)

    # Górna ramka
    print(border)
    # Wiersz nagłówka
    print(draw_row(headers, col_widths))
    # Linia oddzielająca nagłówek od danych
    print(border)

    # Wiersze danych
    for row in rows:
        print(draw_row(row, col_widths))
        print(border)

    return None


# ===========================================================================================
# ---------- FUNKCJA GŁÓWNA (obsługa wejścia i logika programu) ----------
# ===========================================================================================

def main(filepath: str):
    """
    Główna funkcja programu:
    - Wczytuje dane tabelaryczne z pliku CSV lub Excel,
    - Rysuje tabelę w konsoli.
    """

    path = Path(filepath)
    if not path.is_file():
        print(f"Błąd: Plik '{filepath}' nie istnieje.")
        return

    # Wczytujemy dane korzystając z funkcji pomocniczej
    try:
        headers, rows = read_tabular_file(filepath)
    except Exception as e:
        print(f"Błąd podczas wczytywania pliku: {e}")
        return

    if not headers:
        print("Plik wydaje się być pusty lub nie zawiera danych.")
        return

    print("\nWczytana tabela:\n")
    # draw_table SAM policzy szerokości kolumn
    draw_table(headers, rows, max_col_width=30)


if __name__ == "__main__":
    file_path = input("Podaj ścieżkę do pliku CSV lub Excel: ").strip()
    main(file_path)
